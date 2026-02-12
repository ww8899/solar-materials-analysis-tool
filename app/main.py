from __future__ import annotations

import io
import math
from typing import List

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.responses import FileResponse
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from pydantic import BaseModel

app = FastAPI(title="Solar Materials Analysis Tool", version="0.1.0")
app.mount("/static", StaticFiles(directory="static"), name="static")


class PlotDataExportRequest(BaseModel):
    time_ns: List[float]
    avg_intensity: List[float]


@app.get("/")
def index() -> FileResponse:
    return FileResponse("static/index.html")


def _parse_csv_bytes(raw: bytes) -> tuple[List[float], List[float], List[List[float]]]:
    text = raw.decode("utf-8", errors="ignore").strip()
    lines = [line for line in text.splitlines() if line.strip()]
    if len(lines) < 2:
        raise ValueError("CSV must have at least header and one data row")

    header_cells = [c.strip() for c in lines[0].split(",")]
    if len(header_cells) < 3:
        raise ValueError("CSV header is too short")

    try:
        wavelengths = [float(v) for v in header_cells[:-1]]
    except ValueError as exc:
        raise ValueError("Header wavelengths must be numeric") from exc

    times: List[float] = []
    matrix: List[List[float]] = []
    for line in lines[1:]:
        cells = [c.strip() for c in line.split(",")]
        if len(cells) < len(wavelengths) + 1:
            continue
        try:
            intensities = [float(v) for v in cells[: len(wavelengths)]]
            t = float(cells[len(wavelengths)])
        except ValueError:
            continue

        matrix.append(intensities)
        times.append(t)

    if not times:
        raise ValueError("No valid numeric rows found in CSV")
    return wavelengths, times, matrix


def _parse_xlsx_bytes(raw: bytes) -> tuple[List[float], List[float], List[List[float]]]:
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active

    max_col = ws.max_column
    max_row = ws.max_row
    if max_row < 2 or max_col < 3:
        raise ValueError("Excel sheet is too small")

    wavelengths: List[float] = []
    for col in range(1, max_col):
        value = ws.cell(row=1, column=col).value
        if value is None or str(value).strip() == "":
            break
        try:
            wavelengths.append(float(value))
        except (TypeError, ValueError):
            raise ValueError("First row must be numeric wavelengths")

    if len(wavelengths) < 2:
        raise ValueError("Could not parse wavelengths from first row")

    time_col = len(wavelengths) + 1
    times: List[float] = []
    matrix: List[List[float]] = []

    for row in range(2, max_row + 1):
        row_values: List[float] = []
        valid_row = True
        for col in range(1, len(wavelengths) + 1):
            value = ws.cell(row=row, column=col).value
            if value is None:
                valid_row = False
                break
            try:
                row_values.append(float(value))
            except (TypeError, ValueError):
                valid_row = False
                break

        if not valid_row:
            continue

        time_value = ws.cell(row=row, column=time_col).value
        try:
            t = float(time_value)
        except (TypeError, ValueError):
            continue

        matrix.append(row_values)
        times.append(t)

    if not times:
        raise ValueError("No valid data rows found")
    return wavelengths, times, matrix


def _parse_uploaded_matrix(upload: UploadFile, raw: bytes) -> tuple[List[float], List[float], List[List[float]]]:
    name = (upload.filename or "").lower()
    if name.endswith(".xlsx"):
        return _parse_xlsx_bytes(raw)
    if name.endswith(".csv"):
        return _parse_csv_bytes(raw)
    raise ValueError("Only .xlsx and .csv are supported")


def _parse_xy_csv_bytes(raw: bytes) -> tuple[List[float], List[float]]:
    text = raw.decode("utf-8", errors="ignore").strip()
    lines = [line for line in text.splitlines() if line.strip()]
    xs: List[float] = []
    ys: List[float] = []

    for line in lines:
        cells = [c.strip() for c in line.split(",")]
        if len(cells) < 2:
            continue
        try:
            x_val = float(cells[0])
            y_val = float(cells[1])
        except ValueError:
            continue
        xs.append(x_val)
        ys.append(y_val)

    if len(xs) < 2:
        raise ValueError("Need at least two numeric data rows in first two CSV columns")
    return xs, ys


def _parse_xy_xlsx_bytes(raw: bytes) -> tuple[List[float], List[float]]:
    wb = load_workbook(io.BytesIO(raw), data_only=True)
    ws = wb.active
    xs: List[float] = []
    ys: List[float] = []

    for row in range(1, ws.max_row + 1):
        x_cell = ws.cell(row=row, column=1).value
        y_cell = ws.cell(row=row, column=2).value
        try:
            x_val = float(x_cell)
            y_val = float(y_cell)
        except (TypeError, ValueError):
            continue
        xs.append(x_val)
        ys.append(y_val)

    if len(xs) < 2:
        raise ValueError("Need at least two numeric data rows in first two Excel columns")
    return xs, ys


def _parse_uploaded_xy(upload: UploadFile, raw: bytes) -> tuple[List[float], List[float]]:
    name = (upload.filename or "").lower()
    if name.endswith(".xlsx"):
        return _parse_xy_xlsx_bytes(raw)
    if name.endswith(".csv"):
        return _parse_xy_csv_bytes(raw)
    raise ValueError("Only .xlsx and .csv are supported")


def _fit_line_ax(xs: np.ndarray, ys: np.ndarray) -> tuple[dict, np.ndarray]:
    denom = float(np.sum(xs**2))
    if denom == 0:
        raise ValueError("Cannot fit y=ax when all x are zero")
    a = float(np.sum(xs * ys) / denom)
    y_fit = a * xs
    return {"a": a}, y_fit


def _fit_quadratic(xs: np.ndarray, ys: np.ndarray) -> tuple[dict, np.ndarray]:
    coeff = np.polyfit(xs, ys, 2)
    a, b, c = float(coeff[0]), float(coeff[1]), float(coeff[2])
    y_fit = a * xs**2 + b * xs + c
    return {"a": a, "b": b, "c": c}, y_fit


def _fit_log_base(xs: np.ndarray, ys: np.ndarray) -> tuple[dict, np.ndarray]:
    if np.any(xs <= 0):
        raise ValueError("All x must be > 0 for y=log_n(x)")
    lx = np.log(xs)
    denom = float(np.sum(lx**2))
    if denom == 0:
        raise ValueError("Cannot fit log model for provided x values")
    k = float(np.sum(ys * lx) / denom)  # y = k ln(x), where k = 1/ln(n)
    if k == 0:
        raise ValueError("Cannot infer base n from data (k=0)")
    n = float(math.exp(1.0 / k))
    y_fit = np.log(xs) / np.log(n)
    return {"n": n}, y_fit


@app.post("/api/analyze-range")
async def analyze_range(
    file: UploadFile = File(...),
    min_wavelength_nm: float = Form(...),
    max_wavelength_nm: float = Form(...),
):
    if min_wavelength_nm > max_wavelength_nm:
        raise HTTPException(status_code=400, detail="min_wavelength_nm must be <= max_wavelength_nm")

    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    try:
        wavelengths, times, matrix = _parse_uploaded_matrix(file, raw)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    selected_indices = [
        i for i, wl in enumerate(wavelengths) if min_wavelength_nm <= wl <= max_wavelength_nm
    ]
    if not selected_indices:
        raise HTTPException(
            status_code=400,
            detail="No wavelength columns in requested range",
        )

    avg_intensity = []
    for row in matrix:
        values = [row[i] for i in selected_indices]
        avg_intensity.append(sum(values) / len(values))

    return {
        "range_nm": [min_wavelength_nm, max_wavelength_nm],
        "selected_wavelength_count": len(selected_indices),
        "time_ns": times,
        "avg_intensity": avg_intensity,
    }


@app.post("/api/curve-fit")
async def curve_fit(
    file: UploadFile = File(...),
    function_type: str = Form(...),
):
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    try:
        x_vals, y_vals = _parse_uploaded_xy(file, raw)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    xs = np.asarray(x_vals, dtype=float)
    ys = np.asarray(y_vals, dtype=float)

    try:
        if function_type == "linear_ax":
            params, y_fit = _fit_line_ax(xs, ys)
        elif function_type == "quadratic":
            params, y_fit = _fit_quadratic(xs, ys)
        elif function_type == "log_n_x":
            params, y_fit = _fit_log_base(xs, ys)
        else:
            raise HTTPException(status_code=400, detail="Unsupported function type")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    return {
        "function_type": function_type,
        "x_data": x_vals,
        "y_data": y_vals,
        "y_fit": y_fit.tolist(),
        "parameters": params,
    }


@app.post("/api/curve-data")
async def curve_data(file: UploadFile = File(...)):
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Uploaded file is empty")

    try:
        x_vals, y_vals = _parse_uploaded_xy(file, raw)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    return {"x_data": x_vals, "y_data": y_vals}


@app.get("/health")
def health() -> dict:
    return {"status": "ok"}


@app.post("/api/export-plot-data")
def export_plot_data(payload: PlotDataExportRequest):
    if not payload.time_ns or not payload.avg_intensity:
        raise HTTPException(status_code=400, detail="time_ns and avg_intensity must not be empty")
    if len(payload.time_ns) != len(payload.avg_intensity):
        raise HTTPException(status_code=400, detail="time_ns and avg_intensity length mismatch")

    wb = Workbook()
    ws = wb.active
    ws.title = "plot_data"
    ws.append(["time_ns", "avg_intensity"])

    for t, y in zip(payload.time_ns, payload.avg_intensity):
        ws.append([float(t), float(y)])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    headers = {"Content-Disposition": 'attachment; filename="plot_data.xlsx"'}
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
